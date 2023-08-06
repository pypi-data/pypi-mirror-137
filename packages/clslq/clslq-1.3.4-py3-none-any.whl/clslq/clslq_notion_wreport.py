import datetime
import string
import traceback
import pandas

from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from clspy.log import Logger

from .templates import weekreport
from .clslq_notion_report import Report

clslog = Logger().log


class WeekReport(Report):
    def week_belongs(self, sdate, edate):
        s = sdate - datetime.timedelta(days=1)
        e = edate + datetime.timedelta(days=1)
        clslog.info("[{} {}] {}".format(s, e, self.datetime_now))
        if self.datetime_now >= (sdate - datetime.timedelta(
                days=2)) and self.datetime_now <= (edate +
                                                   datetime.timedelta(days=2)):
            return True
        else:
            return False

    def set_worksheet_head(self, head):
        """Set sheet head

        Args:
            head (worksheet head): string
        """
        self.sht.merge_cells('A1:F1')

        self.sht['A1'] = u"本周工作情况" + head.replace('-', "")
        self.sht['A1'].alignment = Alignment(horizontal='center',
                                             vertical='center')
        self.sht['A1'].font = Font(color=colors.BLACK, b=True, size=14)
        self.sht['A1'].fill = PatternFill("solid", fgColor="00FF8080")
        self.sht.row_dimensions[1].height = 20

    def set_worksheet_title(self):
        u"""openpyxl 不支持按列设置样式
        """
        self.sht['A2'] = u"分类"
        self.sht['B2'] = u"事项"
        self.sht['C2'] = u"进展"
        self.sht['D2'] = u"问题"
        self.sht['E2'] = u"解决"
        self.sht['F2'] = u"评审、复盘、总结"
        for col in range(1, 7):
            self.sht.cell(column=col, row=2).font = Font(name=u'微软雅黑',
                                                         bold=True,
                                                         size=12)

    def content_parse_title(self, item):
        result = None
        try:
            title = item[u'名称']['title']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_state(self, item, row):
        result = ''
        try:
            for i in item[u'状态']['multi_select']:
                if i['name'] == u'进行中':
                    for col in range(1, 7):
                        self.sht.cell(column=col, row=row).fill = PatternFill(
                            "solid", fgColor="00CCFFCC")
                if i['name'] == u'遇问题':
                    for col in range(1, 7):
                        self.sht.cell(column=col, row=row).fill = PatternFill(
                            "solid", fgColor="00FFCC99")
                result = "{} {}".format(result, i['name'])
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_type(self, item, row):
        result = ''
        try:
            result = item[u'分类']['select']['name']
            if result == u'工作计划':
                self.sht.cell(column=6, row=row).value = u"下周工作计划"
                for col in range(1, 7):
                    self.sht.cell(column=col, row=row).fill = PatternFill(
                        "solid", fgColor="00CCFFCC")
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_richtext(self, item, text):
        """Parse Notion Column content

        Args:
            item (dict): Notion Column content
            text (str): Unicode string means column title

        Returns:
            str: Cell result
        """
        result = ''
        try:
            title = item[text]['rich_text']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result.replace('\n', ' ')

    def excel_worksheet_fill(self, item, row):
        self.sht['A' + str(row)] = self.content_parse_type(item, row)
        self.sht['B' + str(row)] = self.content_parse_title(item)
        self.sht['C' + str(row)] = self.content_parse_state(item, row)
        self.sht['D' + str(row)] = self.content_parse_richtext(item, u'问题')
        self.sht['E' + str(row)] = self.content_parse_richtext(item, u'解决方法')
        self.sht['F' + str(row)] = self.content_parse_richtext(
            item, u'评审、复盘、总结')

        for c in ('A', 'C'):
            self.sht.column_dimensions[c].width = 10
            for cell in self.sht[c]:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)
                cell.border = self.default_border
        for c in ('B', 'F', 'D', 'E'):
            self.sht.column_dimensions[c].width = 30
            for cell in self.sht[c]:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)
                cell.border = self.default_border

    def excel_worksheet_dump(self, wbname, database):
        # Change sheet name
        self.sht.title = 'WR'
        self.set_worksheet_head(wbname)
        self.set_worksheet_title()

        row = 3
        for node in database['results']:
            item = node['properties']
            self.excel_worksheet_fill(item, row)
            row = row + 1
        self.wb.save(filename=wbname + '.xlsx')
        self.wb.close()

    def pandas_df_fill(self, database):
        _type = []
        _title = []
        _state = []
        _problem = []
        _solve = []
        _conclusion = []
        for node in database['results']:
            item = node['properties']
            _type.append(self.content_parse_type(item, 0))
            _title.append(self.content_parse_title(item))
            _state.append(self.content_parse_state(item, 0))
            _problem.append(self.content_parse_richtext(item, u'问题'))
            _solve.append(self.content_parse_richtext(item, u'解决方法'))
            _conclusion.append(self.content_parse_richtext(item, u'评审、复盘、总结'))
        data = {
            u'分类': pandas.Series(_type, index=range(len(_type))),
            u'事项': pandas.Series(_title, index=range(len(_title))),
            u'进展': pandas.Series(_state, index=range(len(_state))),
            u'问题': pandas.Series(_problem, index=range(len(_problem))),
            u'解决': pandas.Series(_solve, index=range(len(_solve))),
            u'评审复盘总结备注': pandas.Series(_conclusion,
                                       index=range(len(_conclusion)))
        }
        self.df = pandas.DataFrame(data)
        return self.df

    def render_html(self, clsconfig, title, database):
        """Render html form template

        Note that some of the email display methods only support inline-css style,
        This method support inline-css for table headers and rows.

        Args:
            title (str): Title of html and email subject
            df (object): Pandas DataFrame object
        """
        clslog.info("Render[Inline-CSS] html for {}".format(title))
        with open(self.wbname + '.html', encoding='utf-8', mode='w') as f:
            table = str('')
            summary = str('')
            plan = str('')
            i = 0
            j = 0

            task_paln_template = """
                <tr height="19" style="height:14.0pt;background: {color}">
                    <td style="border: 0.5pt solid #cfcfcf; vertical-align: middle; text-align: center;">{type}</td>
                    <td style="border: 0.5pt solid #cfcfcf;">{title}</td>
                    <td style="border: 0.5pt solid #cfcfcf; vertical-align: middle; text-align: center;">{state}</td>
                    <td style="border: 0.5pt solid #cfcfcf;">{problem}</td>
                    <td style="border: 0.5pt solid #cfcfcf;">{solve}</td>
                </tr>
            """
            summary_template = """
                <tr>
                    <td style="padding: 10px; background-color: rgba(204, 204, 204, 0.1)">
                    <span style="font-size: 16px; color: #81e4c3">●</span>&nbsp;
                    <span>
                        <span style="border-bottom: 1px dashed rgb(204, 204, 204); position: relative;">{summary}</span>
                    </span>
                    </td>
                </tr>
            """
            for node in database['results']:
                item = node['properties']

                def bgcolor(i):
                    return '#F7F7F7' if i % 2 == 0 else '#fff'

                if self.content_parse_type(item, 0) == u'工作计划':
                    plan += task_paln_template.format(
                        **{
                            'type': str(i + 1),
                            'title': self.content_parse_title(item),
                            'state': "-",
                            'problem': self.content_parse_richtext(
                                item, u'问题'),
                            'solve': self.content_parse_richtext(
                                item, u'解决方法'),
                            'color': bgcolor(i)
                        })
                    i = i + 1
                else:
                    table += task_paln_template.format(
                        **{
                            'type': self.content_parse_type(item, 0),
                            'title': self.content_parse_title(item),
                            'state': self.content_parse_state(item, 0),
                            'problem': self.content_parse_richtext(
                                item, u'问题'),
                            'solve': self.content_parse_richtext(
                                item, u'解决方法'),
                            'color': bgcolor(j)
                        })
                    j = j + 1
                summarize_item = self.content_parse_richtext(item, u'评审、复盘、总结')
                if len(summarize_item):
                    summary += summary_template.format(
                        **{'summary': summarize_item})
            html = string.Template(weekreport.wr_template)

            f.write(
                html.safe_substitute({
                    "title": title,
                    "table": table,
                    "plan": plan,
                    "user": clsconfig.get('user'),
                    "department": clsconfig.get('department'),
                    "summarize": summary
                }))
