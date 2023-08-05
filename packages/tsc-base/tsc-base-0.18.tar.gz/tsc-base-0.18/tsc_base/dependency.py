import os
try:
    from .func import dict_to_pair, pair_to_dict, merge_dict, get
except:
    from func import dict_to_pair, pair_to_dict, merge_dict, get


class Excel:
    @staticmethod
    def get_excel_th(th_D, horizontal=True, start_r=0, start_c=0, rectangle='$rectangle', sort_key=None, wild_card='$*v', max_deep=None):
        """递归将dict转换为具有单元格位置坐标的excel表头, 用于 xlwt 等, 行列值从0开始

        Args:
            th_D (dict): 会把所有key拿出来作为表头, 然后value中增加 {rectangle:(开始行,结束行,开始列,结束列)}, 注意会覆盖原始value
            horizontal (bool, optional): 是否横向展开表头
            start_r (int, optional): 表头占据的开始行, 通常xlwt从0开始, openpyxl从1开始
            start_c (int, optional): 表头占据的开始列
            rectangle (str, optional): 作为存储每个单元格的4角坐标的key, 也用于sort_key中的func, 不能与th_D中任意key一样
            sort_key (dict, optional): key:{rectangle:{'f':f(k,v),'r':bool},..}; 用于排序同一个字典中的k, 默认不排序
                f: 函数f输入key下面value中的k和v, 返回排序值, k会按照排序值排序, 例如 lambda t:t[0]
                r (bool, optional): 在 f 存在的情况下, 是否倒序排序, 不存在默认False
            wild_card (str, optional): 用于 sort_key 中匹配key的通配符, 如果在sort_key中找不到key则进入wild_card中的value, 不能与th_D中任意key一样
            max_deep (int, optional): 用于递归确定深度, 不能修改

        Returns:
            c, max_deep: int,int; 总占据列宽, 总占据行数
        """
        if max_deep is None:
            x = list(dict_to_pair(th_D))
            for i in x:
                assert len({rectangle, wild_card} & set(i[0])) == 0, 'th_D 中存在 rectangle/wild_card 标记, 请进行修改:' + str(i)
            max_deep = max([len(i[0]) for i in x])
        sort_key = {} if sort_key is None else sort_key
        if rectangle in sort_key:  # 排序
            par = sorted(th_D.items(), key=sort_key[rectangle]['f'], reverse=bool(sort_key[rectangle].get('r')))
        else:
            par = th_D.items()
        c = 0  # 深度优先得到上层的列宽
        for k, v in par:
            start_c_last = start_c + c  # 随着循环, 列的起始点在变化, 行的起始点不变
            if isinstance(v, dict):  # 随着递归, 行每次向下移动一行
                r = 1  # 下面还要分之则只能占一行
                c += Excel.get_excel_th(v, horizontal, start_r+1, start_c_last, rectangle,
                                        sort_key.get(k) or sort_key.get(wild_card), wild_card, max_deep-1)[0]
            else:
                r = max_deep
                c += 1  # 只占一列
                th_D[k] = v = {}
            v[rectangle] = (start_r, start_r+r-1, start_c_last, start_c+c-1)
            v[rectangle] = v[rectangle] if horizontal else [*v[rectangle][2:], *v[rectangle][:2]]
        return c, max_deep

    @staticmethod
    def get_excel_table(doc_L, ignore_th_f=lambda t: t, td_f=lambda t: t[1], horizontal=True, rectangle='$rectangle', **kw):
        """将多个dict转换为excel表格的单元格坐标, 用于生成excel表格

        Args:
            doc_L (list): [{..},..]; 被转换的dict列表
            ignore_th_f (func, optional): 输入 (key_L,value) 返回 (key_L,value) or None, None表示丢弃key_L这个表头
                用于修剪层次过深的表头, 这时候可能需要 td_f 优化dict在单元格中的展示形式
            td_f (func, optional): 用于优化单元格中的值, value是func, 输入 (key_L,value) 返回优化的展示 value
                小心处理每一种值格式, 因为 value 不能是 dict,list,tuple,set 等类型, 否则可能导致 excel 写入出错
            horizontal (bool, optional): 见 get_excel_th
            rectangle (str, optional): 见 get_excel_th
            **kw: 其他参数见 get_excel_th

        Returns:
            dict, list, list: th_D 的一个例子:
                {'inf_jcr': {'$rectangle': (0, 0, 1, 33),
                'Open Access': {'$rectangle': (1, 5, 29, 29)},
                '期刊分区': {'$rectangle': (1, 1, 1, 27),
                        'JCR分区': {'$rectangle': (2, 5, 27, 27)},}}
        """
        # 表头
        pair = []
        for p in dict_to_pair(merge_dict(doc_L)):
            p = ignore_th_f(p)
            if p is not None:
                pair.append(p)
        th_D = pair_to_dict(pair)
        Excel.get_excel_th(th_D, horizontal=horizontal, rectangle=rectangle, **kw)
        th_L = [(i[1], i[0][-2]) for i in dict_to_pair(th_D)]  # 与 td_L 格式一致
        # 表中值
        td_L = []  # [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 与doc_L顺序一致
        for no, doc in enumerate(doc_L):
            coor_v_D = {}  # {(开始行,结束行,开始列,结束列):单元格值,..}
            for p in dict_to_pair(doc):
                p = ignore_th_f(p)  # (key_L,value); 用于获取行列位置, 位置靠前了值也会变
                # 检查这个值是否满足要求
                if p is None:
                    continue
                th = get(p[0], th_D)  # 对应的表头坐标
                if len(th) > 1:  # 没有到叶子结点, 说明doc存在大段的空
                    continue
                # 获取坐标
                _, r, _, c = th[rectangle]  # 右下角单元格坐标
                r, c = (r + no + 1, c) if horizontal else (r, c + no + 1)
                coor = (r, r, c, c)  # (开始行,结束行,开始列,结束列)
                # 保存坐标与单元格值
                if coor not in coor_v_D:  # 防止重复
                    coor_v_D[coor] = td_f((p[0], get(p[0], doc)))
            td_L.append(list(coor_v_D.items()))
        return th_D, th_L, td_L

    @staticmethod
    def excel_add_sheet(workbook, name, th_L, td_L, index=0, save_path=None):
        """使用行列坐标 写入一页 excel 表格

        Args:
            workbook (obj): openpyxl.Workbook() or xlwt.Workbook(encoding='utf8')
            name (str): sheet 名称
            th_L (list): [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 行列编号从0开始, 0行就是第一行
            td_L (list): [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 行列编号从0开始, 0行就是第一行
            index (int, optional): 只用于 openpyxl, 表示插入的 sheet 的位置, xlwt 只是追加
            save_path (str, optional): 写入文件的路径, 会自动添加后缀名

        Returns:
            workbook
        """
        import xlwt  # xlwt>=1.3.0
        import openpyxl  # openpyxl>=3.0.9
        from tqdm import tqdm

        if isinstance(workbook, openpyxl.workbook.workbook.Workbook):  # # openpyxl - xlsx (打开效率更高)
            worksheet = workbook.create_sheet(name, index)
            # 写表头
            for coor, v in th_L:
                coor = [i+1 for i in coor]
                worksheet.merge_cells(start_row=coor[0], start_column=coor[2], end_row=coor[1], end_column=coor[3])
                cell = worksheet.cell(coor[0], coor[2])
                cell.value = v
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                cell.font = openpyxl.styles.Font(bold=True)
            # 单元格值
            for i in tqdm(td_L, f'{name}-写入表格(openpyxl)'):
                for coor, v in i:
                    cell = worksheet.cell(coor[0]+1, coor[2]+1)
                    cell.value = v
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            save_path = save_path + '.xlsx' if save_path else None
        else:  # xlwt - xls (写入快6,7倍)
            worksheet = workbook.add_sheet(name)
            # 单元格值
            style = xlwt.XFStyle()
            alignment = xlwt.Alignment()
            alignment.horz = xlwt.Alignment.HORZ_CENTER
            alignment.vert = xlwt.Alignment.VERT_CENTER
            style.alignment = alignment
            for i in tqdm(td_L, f'{name}-写入表格(xlwt)'):
                for coor, v in i:
                    worksheet.write_merge(*coor, v, style)
            # 写表头
            font = xlwt.Font()
            font.bold = True
            style.font = font
            for coor, v in th_L:
                worksheet.write_merge(*coor, v, style)
            save_path = save_path + '.xls' if save_path else None
        if save_path:
            if not os.path.exists(os.path.dirname(save_path)):
                os.makedirs(os.path.dirname(save_path))
            workbook.save(save_path)
        return worksheet


if __name__ == '__main__':
    # Excel
    print('=' * 10, 'Excel')
    doc_L = [{
        'a1': {'b1': 1, 'b2': 2, 'b3': {'c1': 1, 'c2': 2}},
        'a2': 123,
    } for i in range(10)]
    th_D, th_L, td_L = Excel.get_excel_table(doc_L)
    print('th_D:', th_D)
    import xlwt
    import openpyxl
    Excel.excel_add_sheet(openpyxl.Workbook(), 'test', th_L, td_L, 0, 'test/openpyxl')
    Excel.excel_add_sheet(xlwt.Workbook(encoding='utf8'), 'test', th_L, td_L, 0, 'test/xlwt')
